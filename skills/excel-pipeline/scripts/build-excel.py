"""
Build Excel from JSONL + optional formula registry + issues + highlight rules.
Usage: python build-excel.py <data.jsonl> [--output out.xlsx] [--formulas formula-registry.json]
       [--issues issues.json] [--highlights highlight-rules.json] [--sources sources.json]
"""
import json, sys, argparse, re
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from collections import defaultdict

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
YELLOW = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
RED = PatternFill(start_color="FFB3B3", end_color="FFB3B3", fill_type="solid")
GREEN = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
ISSUE_HIGH = PatternFill(start_color="FFB3B3", end_color="FFB3B3", fill_type="solid")
ISSUE_MED = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
ISSUE_LOW = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")

COLOR_MAP = {"FFF2CC": YELLOW, "FFB3B3": RED, "D9EAD3": GREEN}


def load_jsonl(path):
    records = []
    with open(path, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if line:
                records.append(json.loads(line))
    return records


def detect_fields(records):
    """Detect all fields preserving insertion order."""
    fields = []
    seen = set()
    for rec in records:
        for k in rec:
            if k not in seen:
                fields.append(k)
                seen.add(k)
    return fields


def apply_formula(formula_template, row_num, field_to_col):
    """Replace placeholders in formula template."""
    result = formula_template
    # {A}, {B}, etc -> A2, B2
    for letter in re.findall(r'\{([A-Z]+)\}', result):
        result = result.replace(f'{{{letter}}}', f'{letter}{row_num}')
    # {row} -> row number
    result = result.replace('{row}', str(row_num))
    # {col:field_name} -> column letter
    for match in re.findall(r'\{col:(\w+)\}', result):
        if match in field_to_col:
            col_letter = get_column_letter(field_to_col[match])
            result = result.replace(f'{{col:{match}}}', f'{col_letter}{row_num}')
    return result


def build(data_path, output_path, formulas_path=None, issues_path=None,
          highlights_path=None, sources_path=None):
    records = load_jsonl(data_path)
    if not records:
        print("Empty JSONL")
        return

    fields = detect_fields(records)
    field_to_col = {f: i + 1 for i, f in enumerate(fields)}

    # Load optional configs
    formulas = []
    if formulas_path:
        with open(formulas_path, "r", encoding="utf-8") as f:
            formulas = json.load(f)

    highlight_rules = []
    if highlights_path:
        with open(highlights_path, "r", encoding="utf-8") as f:
            highlight_rules = json.load(f)

    issues = []
    if issues_path:
        with open(issues_path, "r", encoding="utf-8") as f:
            issues = json.load(f)

    sources = []
    if sources_path:
        with open(sources_path, "r", encoding="utf-8") as f:
            sources = json.load(f)

    wb = openpyxl.Workbook()

    # ======= SHEET 1: Main Data =======
    ws = wb.active
    ws.title = "Data"

    # Determine extra formula columns
    formula_cols = {}
    next_col = len(fields) + 1
    for fm in formulas:
        if fm.get("sheet") in ("Data", "Main", None):
            col_name = fm["column"]
            if col_name not in field_to_col:
                field_to_col[col_name] = next_col
                formula_cols[col_name] = fm
                next_col += 1

    all_cols = list(field_to_col.items())

    # Header
    for field, col_idx in all_cols:
        cell = ws.cell(row=1, column=col_idx, value=field)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        # Auto width estimate
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(len(field) + 4, 10), 40)

    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    # Data rows
    for row_idx, rec in enumerate(records, 2):
        for field, col_idx in all_cols:
            if field in formula_cols:
                # Formula column
                fm = formula_cols[field]
                formula = apply_formula(fm["formula"], row_idx, field_to_col)
                cell = ws.cell(row=row_idx, column=col_idx, value=formula)
            else:
                v = rec.get(field, "")
                if v is None:
                    v = ""
                cell = ws.cell(row=row_idx, column=col_idx, value=v)

            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.font = Font(size=9)

        # Apply highlight rules
        for rule in highlight_rules:
            condition = rule.get("condition")
            color = COLOR_MAP.get(rule.get("color", "FFF2CC"), YELLOW)

            if condition == "field_null":
                for field in rule.get("fields", []):
                    if field in field_to_col:
                        v = rec.get(field)
                        if v is None or v == "":
                            ws.cell(row=row_idx, column=field_to_col[field]).fill = color

            elif condition == "field_value":
                field = rule.get("field", "")
                values = rule.get("values", [])
                if field in field_to_col and rec.get(field) in values:
                    ws.cell(row=row_idx, column=field_to_col[field]).fill = color

    ws.auto_filter.ref = ws.dimensions

    # ======= SHEET 2: Lookup sheets (auto-detect) =======
    # Find fields with limited unique values (< 200) for lookup
    for field in fields:
        unique_vals = sorted(set(str(rec.get(field, "")) for rec in records if rec.get(field)))
        if 2 <= len(unique_vals) <= 200 and field.endswith(("_code", "_method", "_partner")):
            ws_lookup = wb.create_sheet(f"Lookup_{field}")
            ws_lookup.cell(row=1, column=1, value=field).fill = HEADER_FILL
            ws_lookup.cell(row=1, column=1).font = HEADER_FONT
            ws_lookup.column_dimensions["A"].width = 30
            for i, val in enumerate(unique_vals, 2):
                ws_lookup.cell(row=i, column=1, value=val).font = Font(size=10)
            ws_lookup.freeze_panes = "A2"

    # ======= SHEET: Issues =======
    if issues:
        ws_issues = wb.create_sheet("Issues")
        issue_fields = ["priority", "psp", "market", "channel", "field", "issue", "detail", "status"]
        widths = [10, 18, 18, 25, 25, 30, 50, 30]

        for col_idx, (field, w) in enumerate(zip(issue_fields, widths), 1):
            cell = ws_issues.cell(row=1, column=col_idx, value=field)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            ws_issues.column_dimensions[get_column_letter(col_idx)].width = w

        ws_issues.freeze_panes = "A2"
        for row_idx, issue in enumerate(issues, 2):
            prio = issue.get("priority", "")
            fill = ISSUE_HIGH if prio == "High" else (ISSUE_MED if prio == "Medium" else ISSUE_LOW)
            for col_idx, field in enumerate(issue_fields, 1):
                cell = ws_issues.cell(row=row_idx, column=col_idx, value=issue.get(field, ""))
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.font = Font(size=9)
                if col_idx == 1:
                    cell.fill = fill
                    cell.font = Font(size=9, bold=True)
        ws_issues.auto_filter.ref = ws_issues.dimensions

    # ======= SHEET: Sources =======
    if sources:
        ws_src = wb.create_sheet("Sources")
        src_fields = list(sources[0].keys()) if sources else ["source", "url", "notes"]
        for col_idx, field in enumerate(src_fields, 1):
            cell = ws_src.cell(row=1, column=col_idx, value=field)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            ws_src.column_dimensions[get_column_letter(col_idx)].width = 40
        for row_idx, src in enumerate(sources, 2):
            for col_idx, field in enumerate(src_fields, 1):
                cell = ws_src.cell(row=row_idx, column=col_idx, value=src.get(field, ""))
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.font = Font(size=10)

    # Save
    wb.save(output_path)
    print(f"Saved: {output_path}")
    print(f"Sheets: {wb.sheetnames}")
    print(f"Data: {len(records)} rows, {len(all_cols)} columns")
    if formulas:
        print(f"Formulas: {len(formula_cols)} columns")
    if issues:
        print(f"Issues: {len(issues)}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Build Excel from JSONL")
    parser.add_argument("input", help="Input JSONL file")
    parser.add_argument("--output", "-o", help="Output Excel file")
    parser.add_argument("--formulas", help="Formula registry JSON")
    parser.add_argument("--issues", help="Issues JSON")
    parser.add_argument("--highlights", help="Highlight rules JSON")
    parser.add_argument("--sources", help="Sources JSON")
    args = parser.parse_args()

    output = args.output or args.input.replace(".jsonl", ".xlsx")
    build(args.input, output, args.formulas, args.issues, args.highlights, args.sources)
